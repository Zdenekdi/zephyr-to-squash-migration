#!/usr/bin/env python3
"""
convert.py
----------
Offline utility to convert Zephyr Scale Excel exports to Squash TM Excel imports.

Usage:
    python convert.py --input zephyr_export.xlsx --output squash_import.xlsx --project "SquashProject"
"""

import argparse
import os
import re
import sys
import openpyxl
from openpyxl.utils import get_column_letter

# Zephyr Scale Excel header search patterns (cleaned alphanumeric lowercase)
# Zephyr Scale může exportovat sloupce různými názvy podle verze a nastavení
# Ověřené názvy z uživatelova exportu: Summary, Description, Test Step, Test Result
ZEPHYR_MAPPINGS = {
    "key":          ["key", "id", "testcasekey", "issuekey", "tcjikey", "tckey", "jiratckey",
                     "issuekey"],
    "name":         ["summary",                                   # ← Zephyr Scale: "Summary"
                     "name", "title", "subject", "tcname",
                     "testcasename", "testname", "casename", "issuename", "testcasetitle"],
    "folder":       ["folder", "folderpath", "path", "tcfolder", "component",
                     "testfolder", "suitepath", "module"],
    "status":       ["status", "state", "tcstatus", "teststatus"],
    "priority":     ["priority", "priorityname", "importance", "tcpriority", "severity"],
    "objective":    ["description",                               # ← Zephyr Scale: "Description"
                     "objective", "details", "tcobjective", "tcdescription",
                     "testdescription", "goal"],
    "precondition": ["precondition", "preconditions", "prerequisites", "prerequisite",
                     "tcprecondition", "setup", "prereq"],
    "step_action":  ["teststep",                                  # ← Zephyr Scale: "Test Step"
                     "testscriptstepbystepstep", "step", "stepdescription", "action",
                     "stepaction", "stepname", "testaction", "steppbystepstep"],
    "step_data":    ["testscriptstepbysteptestdata", "testdata", "data", "stepdata",
                     "inputdata", "parameters"],
    "step_expected":["testresult",                                # ← Zephyr Scale: "Test Result"
                     "testscriptstepbystepexpectedresult", "expectedresult", "expected",
                     "stepexpectedresult", "expectedresults", "result", "expectedoutput"]
}

# Importance mapping (Zephyr -> Squash TM)
IMPORTANCE_MAP = {
    "low": "LOW",
    "medium": "MEDIUM",
    "high": "HIGH",
    "very_high": "VERY_HIGH",
    "critical": "VERY_HIGH",
    "major": "HIGH",
    "minor": "LOW"
}

# Status mapping (Zephyr -> Squash TM)
# Použity jen hodnoty které Squash TM bezpečně přijímá: WORK_IN_PROGRESS a APPROVED
STATUS_MAP = {
    "approved":      "APPROVED",
    "done":          "APPROVED",
    "pass":          "APPROVED",
    "passed":        "APPROVED",
    "draft":         "WORK_IN_PROGRESS",
    "inprogress":    "WORK_IN_PROGRESS",
    "underreview":   "WORK_IN_PROGRESS",
    "underrevision": "WORK_IN_PROGRESS",
    "inreview":      "WORK_IN_PROGRESS",
    "obsolete":      "WORK_IN_PROGRESS",
    "deprecated":    "WORK_IN_PROGRESS",
    "wip":           "WORK_IN_PROGRESS",
}
STATUS_DEFAULT = "WORK_IN_PROGRESS"

def clean_header(val) -> str:
    """Removes all non-alphanumeric characters and lowercases the string."""
    if val is None:
        return ""
    return "".join(c for c in str(val).lower() if c.isalnum())

def clean_wiki_markup(text: str) -> str:
    """Removes Confluence/Jira wiki markup from text.
    
    Handles: image refs, links, bold, headings, panels, code blocks, horizontal rules.
    """
    if not text:
        return text
    t = text

    # Confluence image markup – různé formáty:
    # !image.png!  nebo  !image.png|thumbnail!  nebo  !image.png|width=500,alt="..."!
    t = re.sub(r'![^\n!]+\.(?:png|jpg|jpeg|gif|svg|bmp|webp)[^\n!]*!', '', t, flags=re.IGNORECASE)
    # |image-xyz.png|width=658,alt="..."|  (celý token od | do | včetně příloh za jménem)
    t = re.sub(r'\|[^\|\n]*\.(?:png|jpg|jpeg|gif|svg|bmp|webp)[^\|\n]*\|(?:[^\|\n]*\|)?', '', t, flags=re.IGNORECASE)

    # Confluence link: [text|http://url] nebo [http://url]
    t = re.sub(r'\[([^|\]]+)\|https?://[^\]]+\]', r'\1', t)   # [text|url] → text
    t = re.sub(r'\[https?://[^\]]+\]', '', t)                   # [url] → odstraň

    # Confluence bold/italic: *text* → text,  _text_ → text
    t = re.sub(r'\*([^*\n]+)\*', r'\1', t)
    t = re.sub(r'_([^_\n]+)_', r'\1', t)

    # Confluence nadpisy: h1. h2. ...
    t = re.sub(r'^h[1-6]\.\s*', '', t, flags=re.MULTILINE)

    # Confluence code: {code}...{code} nebo {noformat}...{noformat}
    t = re.sub(r'\{code[^}]*\}.*?\{code\}', '[kód]', t, flags=re.DOTALL)
    t = re.sub(r'\{noformat[^}]*\}.*?\{noformat\}', '[text]', t, flags=re.DOTALL)

    # Confluence panely a makra: {panel:...} {info} {note} {warning}
    t = re.sub(r'\{[a-zA-Z][^}]*\}', '', t)

    # Vícenásobné prázdné řádky → jeden
    t = re.sub(r'\n{3,}', '\n\n', t)

    return t.strip()


def clean_html(text: str) -> str:
    """Cleans Confluence wiki markup and converts plain text to HTML paragraphs."""
    if not text:
        return ""
    text_str = str(text).strip()

    # Nejdřív odstraň Confluence wiki markup
    text_str = clean_wiki_markup(text_str)

    # Pokud je to čisté HTML, vrať beze změny
    if text_str.startswith("<") and text_str.endswith(">"):
        return text_str

    # Převeď plain text na HTML odstavce
    paragraphs = text_str.split("\n\n")
    html_p = []
    for p in paragraphs:
        if p.strip():
            p_formatted = p.strip().replace("\n", "<br>")
            html_p.append(f"<p>{p_formatted}</p>")
    return "".join(html_p) if html_p else f"<p>{text_str}</p>"

def map_headers(headers: list) -> dict:
    """Maps header values to target fields based on best substring match."""
    mapping = {}
    cleaned_headers = [clean_header(h) for h in headers]
    
    # Priority 1: Exact matches
    for target, keywords in ZEPHYR_MAPPINGS.items():
        for i, cleaned in enumerate(cleaned_headers):
            if cleaned in keywords:
                mapping[target] = i
                
    # Priority 2: Substring matches for unmatched targets
    for target, keywords in ZEPHYR_MAPPINGS.items():
        if target in mapping:
            continue
        for i, cleaned in enumerate(cleaned_headers):
            if i in mapping.values():
                continue
            for kw in keywords:
                if len(kw) > 3 and (kw in cleaned or cleaned in kw):
                    mapping[target] = i
                    break
            if target in mapping:
                break
                
    return mapping

def sanitize_path(path: str) -> str:
    """Cleans folder path, ensuring forward slashes and no duplicate slashes."""
    if not path:
        return ""
    path_str = str(path).strip().replace("\\", "/")
    # Remove leading/trailing slashes for easier split
    path_str = path_str.strip("/")
    parts = [p.strip() for p in path_str.split("/") if p.strip()]
    
    # Squash TM path escape rule: escape any '/' inside folder names
    # Note: since we split by '/', parts won't contain '/' unless originally escaped.
    # We join them back with '/'.
    return "/".join(parts)

def parse_zephyr_excel(file_path: str) -> list[dict]:
    """Parses Zephyr Excel file into structured test case list."""
    if not os.path.exists(file_path):
        print(f"Error: Vstupní soubor '{file_path}' neexistuje.")
        sys.exit(1)
        
    print(f"Načítám soubor: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Zobraz všechny listy v souboru
    print(f"\nListy v souboru: {wb.sheetnames}")
    sheet = wb.active
    print(f"Aktivní list: '{sheet.title}' ({sheet.max_row} řádků, {sheet.max_column} sloupců)")

    # Zobraz RAW záhlaví (přesně jak jsou v souboru)
    headers = [cell.value for cell in sheet[1]]
    print("\nRaw záhlaví sloupců (řádek 1):")
    for i, h in enumerate(headers):
        print(f"  Sloupec {i+1}: {repr(h)}")

    mapping = map_headers(headers)

    print("\nDetekované mapování sloupců:")
    for target, idx in mapping.items():
        col_letter = get_column_letter(idx + 1)
        print(f"  - {target:15} -> Sloupec {col_letter} ({repr(headers[idx])})")

    # Zobraz nemapované sloupce
    mapped_idxs = set(mapping.values())
    unmapped = [(i, h) for i, h in enumerate(headers) if i not in mapped_idxs and h is not None]
    if unmapped:
        print("\nNemapované sloupce (ignorovány):")
        for i, h in unmapped:
            print(f"  Sloupec {i+1}: {repr(h)}")

    # Pokud "name" nenalezeno – varujeme, ale NEPŘERUŠUJEME (použijeme key jako fallback)
    if "name" not in mapping:
        if "key" in mapping:
            print("\nVarování: Sloupec s názvem testu (Summary/Name) nenalezen.")
            print("  Jako název testu bude použit klíč (Issue Key / TC_REFERENCE).")
        else:
            print("\nChyba: Nepodařilo se namapovat ani název (name) ani klíč (key).")
            print("  Ujistěte se, že soubor je Zephyr export (ne již konvertovaný soubor).")
            sys.exit(1)

    test_cases = []
    current_tc = None

    for row_idx in range(2, sheet.max_row + 1):
        row = [sheet.cell(row=row_idx, column=col_idx + 1).value for col_idx in range(len(headers))]

        # Read fields
        tc_key = str(row[mapping["key"]]).strip() if "key" in mapping and row[mapping["key"]] is not None else ""
        # Název: ze sloupce name, nebo fallback na key
        if "name" in mapping and row[mapping["name"]] is not None:
            tc_name = str(row[mapping["name"]]).strip()
        else:
            tc_name = ""  # fallback se provede níže přes tc_display_name
        
        # If we find a new test case key or name, we start a new test case
        is_new_tc = False
        if tc_name or tc_key:
            # Check if this is truly a new test case or just duplicate row of same test case
            if not current_tc:
                is_new_tc = True
            elif tc_key and tc_key != current_tc["key"]:
                is_new_tc = True
            elif tc_name and tc_name != current_tc["name"] and not tc_key:
                is_new_tc = True
                
        if is_new_tc:
            if current_tc:
                test_cases.append(current_tc)

            # Fallback pro název: pokud je prázdný, použijeme klíč (napr. EDAZ-123)
            tc_display_name = tc_name or tc_key or f"TC_{row_idx}"

            current_tc = {
                "key": tc_key,
                "name": tc_display_name,
                "folder": sanitize_path(row[mapping["folder"]]) if "folder" in mapping and row[mapping["folder"]] is not None else "",
                "status": str(row[mapping["status"]]).strip() if "status" in mapping and row[mapping["status"]] is not None else "",
                "priority": str(row[mapping["priority"]]).strip() if "priority" in mapping and row[mapping["priority"]] is not None else "",
                "objective": str(row[mapping["objective"]]).strip() if "objective" in mapping and row[mapping["objective"]] is not None else "",
                "precondition": str(row[mapping["precondition"]]).strip() if "precondition" in mapping and row[mapping["precondition"]] is not None else "",
                "steps": []
            }
            
        # Extract steps for current test case
        step_act = str(row[mapping["step_action"]]).strip() if "step_action" in mapping and row[mapping["step_action"]] is not None else ""
        step_exp = str(row[mapping["step_expected"]]).strip() if "step_expected" in mapping and row[mapping["step_expected"]] is not None else ""
        step_dat = str(row[mapping["step_data"]]).strip() if "step_data" in mapping and row[mapping["step_data"]] is not None else ""
        
        # If there's step data, append it to step action to preserve it
        if step_dat and step_act:
            step_act = f"{step_act}\n\n[TestData: {step_dat}]"
            
        if current_tc and (step_act or step_exp):
            current_tc["steps"].append({
                "action": step_act,
                "expected": step_exp
            })
            
    # Add the last test case
    if current_tc:
        test_cases.append(current_tc)

    wb.close()
    print(f"\nÚspěšně načteno {len(test_cases)} testovacích případů ze Zephyr exportu.")

    # Diagnostika – ukáže prvních 5 TC aby bylo vidět co se parsuje
    print("\n--- DIAGNOSTIKA: prvních 5 testovacích případů ---")
    for tc in test_cases[:5]:
        steps_ok = len(tc['steps'])
        first_step = tc['steps'][0]['action'][:60] if tc['steps'] else '(žádné stepy)'
        print(f"  KEY:    {repr(tc['key'])}")
        print(f"  NAME:   {repr(tc['name'])}")
        print(f"  FOLDER: {repr(tc['folder'])}")
        print(f"  STEPS:  {steps_ok} kroků | 1. krok: {repr(first_step)}")
        print()
    return test_cases

def write_squash_excel(test_cases: list[dict], output_path: str, project_name: str) -> None:
    """Generates Squash TM compatible Excel import file.
    Column names verified against official Squash TM import template.
    """
    print(f"\nVytvářím Squash TM importní soubor: {output_path}")
    wb = openpyxl.Workbook()

    # 1. Sheet TEST_CASES – dle officiálního template
    # TC_WEIGHT (ne TC_IMPORTANCE), TC_PRE_REQUISITE (ne TC_PREREQUISITES)
    ws_tc = wb.active
    ws_tc.title = "TEST_CASES"
    ws_tc.append([
        "ACTION", "TC_PATH", "TC_NUM", "TC_REFERENCE", "TC_NAME",
        "TC_WEIGHT", "TC_STATUS", "TC_DESCRIPTION", "TC_PRE_REQUISITE"
    ])

    # 2. Sheet STEPS – TC_STEP_ACTION, TC_STEP_EXPECTED_RESULT, TC_STEP_NUM
    ws_steps = wb.create_sheet(title="STEPS")
    ws_steps.append([
        "ACTION", "TC_OWNER_PATH", "TC_STEP_NUM",
        "TC_STEP_ACTION", "TC_STEP_EXPECTED_RESULT"
    ])

    # 3. Sheet PARAMETERS – TC_PARAM_DESCRIPTION přidán dle template
    ws_params = wb.create_sheet(title="PARAMETERS")
    ws_params.append(["ACTION", "TC_OWNER_PATH", "TC_PARAM_NAME", "TC_PARAM_DESCRIPTION"])

    # 4. Sheet DATASETS – TC_PARAM_OWNER_PATH přidán dle template
    ws_datasets = wb.create_sheet(title="DATASETS")
    ws_datasets.append([
        "ACTION", "TC_OWNER_PATH", "TC_DATASET_NAME",
        "TC_PARAM_OWNER_PATH", "TC_DATASET_PARAM_NAME", "TC_DATASET_PARAM_VALUE"
    ])

    # 5. Sheet LINK_REQ_TC – bez ACTION a TC_NAME dle template
    ws_links = wb.create_sheet(title="LINK_REQ_TC")
    ws_links.append(["REQ_PATH", "REQ_VERSION_NUM", "TC_PATH"])
    
    # Map and write test cases
    for tc in test_cases:
        # Cesta: /NazevProjektu/Slozka
        # Squash TM NEPŘIJÍMÁ cestu přímo do kořene projektu (napr. /EDAZ)
        # – musí být vždy alespoň jedna podsložka (napr. /EDAZ/Importovane_testy)
        project_clean = project_name.strip("/").strip()
        if tc['folder']:
            squash_path = f"/{project_clean}/{tc['folder']}"
        else:
            squash_path = f"/{project_clean}/Importovane_testy"

        # Status mapping – použij jen bezpečné hodnoty
        cleaned_status = clean_header(tc["status"])
        squash_status = STATUS_MAP.get(cleaned_status, STATUS_DEFAULT)

        # Weight (importance) mapping – TC_WEIGHT dle template
        cleaned_priority = clean_header(tc["priority"])
        squash_weight = IMPORTANCE_MAP.get(cleaned_priority, "MEDIUM")

        # HTML formátování
        description_html = clean_html(tc["objective"])
        precondition_html = clean_html(tc["precondition"])

        # TEST_CASES řádek
        ws_tc.append([
            "C",                  # ACTION
            squash_path,          # TC_PATH
            None,                 # TC_NUM (prázdná buňka, ne "", jinak Squash TM selže při parsování)
            tc["key"],            # TC_REFERENCE
            tc["name"],           # TC_NAME
            squash_weight,        # TC_WEIGHT
            squash_status,        # TC_STATUS
            description_html,     # TC_DESCRIPTION
            precondition_html,    # TC_PRE_REQUISITE
        ])

        # STEPS řádky
        for step_num, step in enumerate(tc["steps"], start=1):
            ws_steps.append([
                "C",                              # ACTION
                squash_path,                      # TC_OWNER_PATH
                step_num,                         # TC_STEP_NUM
                clean_html(step["action"]),       # TC_STEP_ACTION
                clean_html(step["expected"]),     # TC_STEP_EXPECTED_RESULT
            ])

    # Uložení
    wb.save(output_path)
    wb.close()
    print(f"Soubor úspěšně uložen. Obsahuje:")
    print(f"  - {len(test_cases)} řádků v listu TEST_CASES")
    print(f"  - {sum(len(tc['steps']) for tc in test_cases)} řádků v listu STEPS")

def main():
    parser = argparse.ArgumentParser(description="Převede export testů ze Zephyr Scale na Squash TM Excel import.")
    parser.add_argument("-i", "--input", required=True, help="Cesta k exportovanému Excel souboru ze Zephyru (.xlsx).")
    parser.add_argument("-o", "--output", default="squash_import.xlsx", help="Cesta pro uložení Squash TM souboru (výchozí: squash_import.xlsx).")
    parser.add_argument("-p", "--project", default="Imported_Project", help="Název projektu ve Squash TM pro prefix složek (výchozí: Imported_Project).")
    
    args = parser.parse_args()
    
    try:
        test_cases = parse_zephyr_excel(args.input)
        if not test_cases:
            print("Chyba: V souboru nebyly nalezeny žádné testovací případy.")
            sys.exit(1)
        write_squash_excel(test_cases, args.output, args.project)
        print("\nHotovo! Nyní můžete vygenerovaný soubor importovat do Squash TM.")
    except Exception as e:
        print(f"\nNeočekávaná chyba při zpracování: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
